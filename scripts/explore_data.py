"""Quick summary of the raw data under ``data/``.

Reports:
* Number and (min/max/mean peak) of force signals from the Excel file.
* Shape, channel count, node count, and (force_idx, frf_idx, node_id)
  uniqueness for the HDF5 response file.
* Number of (force_idx, node_id) groups and their inner channel layout.

Usage
-----
>>> python -m scripts.explore_data
"""
from __future__ import annotations

from pathlib import Path

import h5py
import numpy as np
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
FORCE_XLSX = ROOT / "data" / "diverse_impact_signals.xlsx"
RESP_H5 = ROOT / "data" / "batch_response_results.h5"


def summarize_forces() -> None:
    print(f"\n=== force file: {FORCE_XLSX.name} ===")
    wb = load_workbook(str(FORCE_XLSX), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    data = np.asarray(rows[1:], dtype=np.float64)
    n_time = data.shape[0]
    n_forces = len(header) - 1
    forces = data[:, 1:].T                           # (N, T)
    peaks = np.abs(forces).max(axis=1)
    print(f"  time steps       : {n_time}")
    print(f"  # force signals  : {n_forces}")
    print(f"  peak |force|     : min={peaks.min():.4g}  max={peaks.max():.4g}  "
          f"mean={peaks.mean():.4g}")
    print(f"  dtype            : {forces.dtype}")


def summarize_responses() -> None:
    print(f"\n=== response file: {RESP_H5.name} ===")
    with h5py.File(RESP_H5, "r") as f:
        print("  keys             :", list(f.keys()))
        resp = f["responses"]
        print(f"  responses.shape  : {resp.shape}   (T, N_cols)")
        force_idx = f["force_signal_indices"][:]
        frf_idx   = f["frf_indices"][:]
        node_ids  = f["node_ids"][:]
        print(f"  # unique force_idx : {np.unique(force_idx).size}")
        print(f"  # unique frf_idx   : {np.unique(frf_idx).size}")
        print(f"  # unique node_ids  : {np.unique(node_ids).size}")
        # uniqueness of (fi, ni, ch) triples
        triples = np.stack([force_idx, frf_idx, node_ids], axis=1)
        uniq = np.unique(triples, axis=0)
        print(f"  (force,frf,node) unique rows: {uniq.shape[0]} "
              f"(expected = total cols = {resp.shape[1]})")
        # group by (force, node) => should always give 3 channels each
        pairs = np.stack([force_idx, node_ids], axis=1)
        upairs = np.unique(pairs, axis=0)
        print(f"  # (force, node) groups: {upairs.shape[0]} "
              f"(expected = {np.unique(force_idx).size} * {np.unique(node_ids).size})")


def main() -> None:
    summarize_forces()
    summarize_responses()


if __name__ == "__main__":
    main()
